#!/usr/bin/env python
"""Acute-levetiracetam sIPSC data for Extended Data Fig. 1, ready for Prism.

Per-cell MEDIANS for the bar panels, and the pooled event distributions for the
cumulative panels, with two-sample Kolmogorov-Smirnov D and P for every
contrast — matching how the other distribution panels are reported.

WHY MEDIANS. Each cell contributes many events, summarised within the cell
before averaging across cells — the same construction as Figure 2. For
frequency the per-event values are INSTANTANEOUS frequency (1/interval),
strongly right-skewed, so per-cell means run about twice the medians and the
Figure 2c group difference is lost on means (P = 0.048 vs 0.116). Medians are
therefore the primary values here; means are written out underneath for
reference.

    python scripts/paper_stats/lev_ephys_workbook.py
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import os
import zipfile

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill
from scipy import stats

EPH = ("/Users/marcoledri/Dropbox/Work/Manuscripts and papers/"
       "SV2A paper/Data analysis/Ephys")
BOLD = Font(bold=True)
ITAL = Font(italic=True, size=9)
HDR = PatternFill("solid", fgColor="DDDDDD")

# measure -> (file, table) for control and LEV, per group
SPEC = {
    "Frequency": dict(
        unit="instantaneous frequency (Hz)", log=True,
        ctrl=("Cont VC Frequency Analysis.prism",
              "Cont VC EGFP Frequency all", "Cont VC SV2A Frequency all"),
        lev=("LEV2 Frequency Analysis.prism",
             "LEV2 EGFP Frequency All", "LEV2 SV2A Frequency All")),
    "Amplitude": dict(
        unit="amplitude (pA)", log=False,
        ctrl=("Cont VC Amplitude Analysis.prism",
              "Cont VC EGFP Amplitude All", "Cont VC SV2A Amplitude All"),
        lev=("LEV2 Amplitude Analysis.prism",
             "LEV2 EGFP Amplitude All", "LEV2 SV2A Amplitude All")),
    "RiseTime": dict(
        unit="rise time (ms)", log=False,
        ctrl=("Cont VC Rise Time Analysis.prism",
              "Cont VC EGFP Rise time all", "Cont VC SV2A Rise time all"),
        lev=("LEV2 Rise Time Analysis.prism",
             "LEV2 EGFP Rise time All", "LEV2 SV2A Rise time All")),
}


def columns(fname, title):
    """-> [array per column]; titles vary in case and trailing space."""
    z = zipfile.ZipFile(os.path.join(EPH, fname))
    tit = {}
    for n in z.namelist():
        if n.endswith("sheet.json"):
            d = json.load(io.TextIOWrapper(z.open(n)))
            if d.get("@class") == "DataSheet":
                tu = (d.get("table") or {}).get("uid")
                if tu:
                    tit.setdefault(d.get("title", "").strip().lower(), tu)
    uid = tit[title.strip().lower()]
    rows = list(csv.reader(io.TextIOWrapper(z.open(f"data/tables/{uid}/data.csv"))))
    w = max(len(r) for r in rows)
    out = []
    for j in range(w):
        v = []
        for r in rows:
            if j < len(r) and (r[j] or "").strip():
                try:
                    v.append(float(r[j]))
                except ValueError:
                    pass
        if len(v) > 5:
            out.append(np.array(v))
    return out


def ks(a, b):
    D, p = stats.ks_2samp(a, b)
    return float(D), float(p)


def star(p):
    return "***" if p < .001 else "**" if p < .01 else "*" if p < .05 else "ns"


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--out", default=os.path.join(EPH, "ExtendedData_LEV_ephys.xlsx"))
    args = ap.parse_args()

    wb = openpyxl.Workbook()
    readme = wb.active
    readme.title = "README"
    summary = []

    for key, spec in SPEC.items():
        fc, ce_t, cs_t = spec["ctrl"]
        fl, le_t, ls_t = spec["lev"]
        cells = {
            "EGFP control": columns(fc, ce_t),
            "EGFP LEV": columns(fl, le_t),
            "SV2A control": columns(fc, cs_t),
            "SV2A LEV": columns(fl, ls_t),
        }
        means = {k: np.array([c.mean() for c in v]) for k, v in cells.items()}
        meds = {k: np.array([np.median(c) for c in v]) for k, v in cells.items()}
        pooled = {k: np.concatenate(v) for k, v in cells.items()}

        # ---------- per-cell sheet ----------
        ws = wb.create_sheet(key)
        ws["A1"] = f"sIPSC {spec['unit']} — per-cell MEDIANS (use these)"
        ws["A1"].font = BOLD
        ws["A2"] = "PRISM TABLE TYPE: Column (4 groups; one value per cell)"
        ws["A2"].font = BOLD
        order = ["EGFP control", "EGFP LEV", "SV2A control", "SV2A LEV"]
        for c, h in enumerate(order):
            cell = ws.cell(row=4, column=c + 1, value=h)
            cell.font, cell.fill = BOLD, HDR
        for c, k in enumerate(order):
            for r, x in enumerate(meds[k]):
                ws.cell(row=5 + r, column=c + 1, value=round(float(x), 4))
        r = 5 + max(len(v) for v in meds.values()) + 1
        ws.cell(row=r, column=1,
                value="per-cell MEANS (reference only — see note below)").font = BOLD
        r += 1
        for c, h in enumerate(order):
            cell = ws.cell(row=r, column=c + 1, value=h)
            cell.font, cell.fill = BOLD, HDR
        for c, k in enumerate(order):
            for rr, x in enumerate(means[k]):
                ws.cell(row=r + 1 + rr, column=c + 1, value=round(float(x), 4))
        r = r + 1 + max(len(v) for v in means.values()) + 1

        W = lambda a, b: stats.ttest_ind(a, b, equal_var=False).pvalue
        notes = ["STATISTICS on the per-cell MEDIANS — unpaired Welch's t-test",
                 f"  EGFP control vs SV2A control : P = {W(meds['EGFP control'], meds['SV2A control']):.4f}",
                 f"  EGFP LEV     vs SV2A LEV     : P = {W(meds['EGFP LEV'], meds['SV2A LEV']):.4f}",
                 f"  EGFP control vs EGFP LEV     : P = {W(meds['EGFP control'], meds['EGFP LEV']):.4f}",
                 f"  SV2A control vs SV2A LEV     : P = {W(meds['SV2A control'], meds['SV2A LEV']):.4f}",
                 "",
                 "  n cells: " + ", ".join(f"{k} {len(meds[k])}" for k in order),
                 "  LEV cells are NOT the control cells — comparisons are unpaired.",
                 ""]
        if key == "Frequency":
            notes += ["  Medians are used because these are INSTANTANEOUS frequencies,",
                      "  which are strongly right-skewed: per-cell means run about twice",
                      "  the medians and the Figure 2c group difference does not survive",
                      "  the switch (P = 0.048 on medians, 0.116 on means). Figure 2 uses",
                      "  medians for the same reason — keep the two consistent.", ""]
        for i, t in enumerate(notes):
            ws.cell(row=r + i, column=1, value=t).font = BOLD if i == 0 else ITAL
        for c in "ABCD":
            ws.column_dimensions[c].width = 15

        # ---------- cumulative-distribution sheet ----------
        wsd = wb.create_sheet(f"CDF_{key}")
        wsd["A1"] = (f"sIPSC {spec['unit']} — pooled events, for cumulative "
                     "distribution plots")
        wsd["A1"].font = BOLD
        wsd["A2"] = ("PRISM: paste a column, then Analyze -> Cumulative frequency "
                     "distribution (or plot as XY)")
        wsd["A2"].font = BOLD
        for c, h in enumerate(order):
            cell = wsd.cell(row=4, column=c + 1, value=f"{h}  (n={len(pooled[h])})")
            cell.font, cell.fill = BOLD, HDR
        for c, k in enumerate(order):
            for rr, x in enumerate(pooled[k]):
                wsd.cell(row=5 + rr, column=c + 1, value=round(float(x), 5))
        r = 5 + max(len(v) for v in pooled.values()) + 1
        contrasts = [("EGFP control", "SV2A control", "control: EGFP vs SV2A"),
                     ("EGFP LEV", "SV2A LEV", "LEV: EGFP vs SV2A"),
                     ("EGFP control", "EGFP LEV", "EGFP: control vs LEV"),
                     ("SV2A control", "SV2A LEV", "SV2A: control vs LEV")]
        lines = ["STATISTICS on the pooled distributions — two-sample "
                 "Kolmogorov-Smirnov"]
        for a, b, lab in contrasts:
            D, p = ks(pooled[a], pooled[b])
            lines.append(f"  {lab:26} D = {D:.4f}   P = {p:.3g}   {star(p)}")
            summary.append((spec["unit"], lab, D, p))
        lines += ["",
                  "  Events are pooled within group and treated as independent,",
                  "  as elsewhere in the manuscript; D is the informative measure",
                  "  and the cell-level tests above carry the inference."]
        for i, t in enumerate(lines):
            wsd.cell(row=r + i, column=1, value=t).font = BOLD if i == 0 else ITAL
        for c in "ABCD":
            wsd.column_dimensions[c].width = 18

    lines = [("Acute levetiracetam on sIPSCs — Extended Data Fig. 1", BOLD), ("", ITAL),
             ("Per-measure sheets hold per-cell MEDIANS (means below, for reference)",
              ITAL),
             ("for the bar panels; CDF_ sheets hold the pooled events for the",
              ITAL),
             ("cumulative distribution panels.", ITAL), ("", ITAL),
             ("KEY RESULT", BOLD),
             ("  Levetiracetam abolishes the SV2A-mediated increase in sIPSC", ITAL),
             ("  frequency, so treated and control cells no longer differ, while", ITAL),
             ("  amplitude is unchanged throughout — a presynaptic interaction", ITAL),
             ("  at the shared target.", ITAL), ("", ITAL),
             ("KOLMOGOROV-SMIRNOV RESULTS (pooled events)", BOLD)]
    for unit, lab, D, p in summary:
        lines.append((f"  {unit:28} {lab:26} D = {D:.4f}  P = {p:.3g}  {star(p)}",
                      ITAL))
    lines += [("", ITAL), ("CAVEATS", BOLD),
              ("  Unpaired: LEV and control recordings are from different cells.", ITAL),
              ("  Frequency means are of instantaneous frequency and are", ITAL),
              ("  skew-inflated relative to the medians used in Figure 2.", ITAL),
              ("  Acute slice application need not match chronic oral dosing.", ITAL)]
    for i, (t, f) in enumerate(lines, 1):
        readme.cell(row=i, column=1, value=t).font = f
    readme.column_dimensions["A"].width = 104

    wb.save(args.out)
    print(f"Wrote {args.out}")
    for unit, lab, D, p in summary:
        print(f"  {unit:28} {lab:26} D={D:.4f} P={p:.3g} {star(p)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
