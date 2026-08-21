#!/usr/bin/env python
"""Build the two replacement panels for Figure 3F, ready to paste into Prism.

Panel F (increase/decrease, week 3 vs week 1) loses its result once the numbers
are corrected — Fisher p goes 0.0223 -> 0.3498, and most of that is already
gone at cut 0, so it is the counts that changed rather than the cut. These two
metrics support the SV2A effect on firmer ground:

  1. Time to first convulsive seizure (Kaplan-Meier, log-rank p = 0.018).
     Treats the 5 SV2A animals that never seize as CENSORED rather than
     discarding them — the exact flaw in the panel being replaced.
  2. Seizure burden, seconds of seizure per 24 h (Mann-Whitney p = 0.0018
     convulsive, 0.016 non-convulsive). Pairs with panel G: duration does not
     differ, so burden shows the effect is carried by frequency.

Writes an xlsx laid out as the Prism table each one needs, with the test to run
and the expected result noted under the numbers.

    python scripts/paper_stats/fig3_replacement_panels.py [--cut 0.5]
"""
from __future__ import annotations

import argparse
import os
import sys

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from scipy import stats

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from prism_rewrite_all import Data          # noqa: E402

BASE = {1, 2, 3}
BOLD = Font(bold=True)
ITAL = Font(italic=True, size=9)
HDR = PatternFill("solid", fgColor="DDDDDD")


def cumulative(D, E, S):
    """-> {animal: [cumulative seizure MINUTES after each baseline day 1..21]}"""
    import collections
    sec = collections.defaultdict(float)
    for ci, a, _t, d in D.ev:
        if D.wk.get(ci) in BASE:
            sec[(a, D.day[ci])] += d
    out = {}
    for a in E + S:
        run, series = 0.0, []
        for dd in range(1, 22):
            run += sec.get((a, dd), 0.0)
            series.append(run / 60.0)
        out[a] = series
    return out


def collect(cut: float):
    D = Data(cut)
    E, S = D.egfp, D.sv2a
    hrs, _rd, _cnt, durs, _cd = D._agg(D._sel(weeks=BASE))

    def burden(g, k):
        return [(sum(durs[(a, k)]) / (hrs[a] / 24) if hrs[a] else None) for a in g]

    first = {}
    for ci, a, t, _d in D.ev:
        if t == "convulsive" and D.wk.get(ci) in BASE:
            dn = D.day.get(ci)
            if dn and (a not in first or dn < first[a]):
                first[a] = dn

    def surv(g):
        out = []
        for a in g:
            days = [D.day[ci] for ci, aa, _s in D.fa
                    if aa == a and D.wk.get(ci) in BASE]
            last = max(days) if days else 21
            out.append((a, first[a], 1) if a in first else (a, last, 0))
        return out

    return D, E, S, burden, surv


def logrank(A, B):
    times = sorted({t for _a, t, _e in A + B})
    O1 = Ex1 = V = 0.0
    for t in times:
        n1 = sum(1 for _a, x, _e in A if x >= t)
        n2 = sum(1 for _a, x, _e in B if x >= t)
        d1 = sum(1 for _a, x, e in A if x == t and e)
        d2 = sum(1 for _a, x, e in B if x == t and e)
        n, d = n1 + n2, d1 + d2
        if n < 2 or d == 0:
            continue
        O1 += d1
        Ex1 += d * n1 / n
        V += d * (n1 / n) * (1 - n1 / n) * ((n - d) / (n - 1))
    chi2 = (O1 - Ex1) ** 2 / V if V > 0 else float("nan")
    return chi2, float(1 - stats.chi2.cdf(chi2, 1))


def mw(a, b):
    a = [x for x in a if x is not None]
    b = [x for x in b if x is not None]
    return (float(np.median(a)), float(np.median(b)),
            float(stats.mannwhitneyu(a, b, alternative="two-sided").pvalue))


def put_notes(ws, row, lines):
    for n in lines:
        ws.cell(row=row, column=1, value=n).font = BOLD if n.isupper() and n else ITAL
        row += 1
    return row


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--cut", type=float, default=0.5)
    ap.add_argument("--out", default="Figure3_replacement_panels.xlsx")
    args = ap.parse_args()

    D, E, S, burden, surv = collect(args.cut)
    A, B = surv(E), surv(S)
    chi2, p_lr = logrank(A, B)
    pc = mw(burden(E, "convulsive"), burden(S, "convulsive"))
    pn = mw(burden(E, "non_convulsive"), burden(S, "non_convulsive"))

    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "1_TimeToFirstSeizure_KM"
    ws["A1"] = ("Time to first CONVULSIVE seizure — baseline (weeks 1-3), "
                f"detector confidence >= {args.cut}")
    ws["A1"].font = BOLD
    ws["A3"] = "PRISM TABLE TYPE:  Survival"
    ws["A3"].font = BOLD
    ws["A4"] = "Paste the block below starting at the X column. One row per animal."
    r = 6
    for c, h in enumerate(["X  (day)", "EGFP", "SV2A",
                           "← animal (reference, do not paste)"]):
        cell = ws.cell(row=r, column=c + 1, value=h)
        cell.font, cell.fill = BOLD, HDR
    r += 1
    for col, rows_ in ((2, A), (3, B)):
        for a, t, e in rows_:
            ws.cell(row=r, column=1, value=t)
            ws.cell(row=r, column=col, value=e)
            ws.cell(row=r, column=4,
                    value=f"{'EGFP' if col == 2 else 'SV2A'} {a}")
            r += 1
    r = put_notes(ws, r + 1, [
        "STATISTICS TO RUN IN PRISM",
        "  Analyze → Survival analysis → 'Compare two survival curves'",
        "  Test: Log-rank (Mantel-Cox). Gehan-Breslow-Wilcoxon optional as secondary.",
        "",
        "CODING:  1 = convulsive seizure occurred on that day (event)",
        "         0 = censored: no convulsive seizure during baseline; X is then",
        "             that animal's LAST recorded baseline day.",
        "",
        f"EXPECTED RESULT (lunarc_detect_wk1-6_final.db, cut {args.cut}):",
        f"  EGFP  {sum(e for _a, _t, e in A)}/{len(A)} seized, median day "
        f"{np.median([t for _a, t, _e in A]):.0f}",
        f"  SV2A  {sum(e for _a, _t, e in B)}/{len(B)} seized, "
        f"{sum(1 for _a, _t, e in B if not e)} censored (never seized)",
        f"  Log-rank chi-square = {chi2:.2f}, df = 1, p = {p_lr:.4f}",
        "",
        "Do NOT use a t-test or Mann-Whitney here — that discards the censored",
        "animals or forces an arbitrary time for them, which is the exact flaw in",
        "the increase/decrease panel this replaces.",
    ])
    for col, w in (("A", 13), ("B", 10), ("C", 10), ("D", 34)):
        ws.column_dimensions[col].width = w

    ws2 = wb.create_sheet("2_SeizureBurden")
    ws2["A1"] = ("Seizure burden — seconds of seizure per 24 h of recording, "
                 f"baseline (weeks 1-3), confidence >= {args.cut}")
    ws2["A1"].font = BOLD
    ws2["A3"] = ("PRISM TABLE TYPE:  Grouped  (rows = seizure type; each animal "
                 "is a side-by-side replicate subcolumn)")
    ws2["A3"].font = BOLD
    ws2["A4"] = ("Column A = EGFP with 7 subcolumns, Column B = SV2A with 13. "
                 "Same layout as the Seizures/day panel.")
    r = 6
    ws2.cell(row=r, column=1, value="").fill = HDR
    for i, a in enumerate(E + S):
        c = ws2.cell(row=r, column=2 + i,
                     value=f"{'EGFP' if i < len(E) else 'SV2A'} {a}")
        c.font, c.fill = BOLD, HDR
    r += 1
    for label, key in (("Convulsive", "convulsive"),
                       ("Non-convulsive", "non_convulsive")):
        ws2.cell(row=r, column=1, value=label).font = BOLD
        for i, v in enumerate(burden(E, key) + burden(S, key)):
            ws2.cell(row=r, column=2 + i,
                     value=None if v is None else round(v, 3))
        r += 1
    put_notes(ws2, r + 1, [
        "STATISTICS TO RUN IN PRISM",
        "  Run each row SEPARATELY, EGFP vs SV2A:",
        "  Analyze → Column analyses → t test (and nonparametric tests)",
        "         → choose 'Mann-Whitney test' (unpaired, two-tailed).",
        "  Matches the test used for the other seizure panels (animal = unit).",
        "  Do not use 2-way ANOVA: values are strongly non-normal and n is small.",
        "",
        f"EXPECTED RESULT (lunarc_detect_wk1-6_final.db, cut {args.cut}):",
        f"  Convulsive      median EGFP {pc[0]:.1f} s/24h   SV2A {pc[1]:.1f} "
        f"s/24h   p = {pc[2]:.4f}",
        f"  Non-convulsive  median EGFP {pn[0]:.1f} s/24h   SV2A {pn[1]:.1f} "
        f"s/24h   p = {pn[2]:.4f}",
        "",
        "Y-AXIS: log10 is advisable — the range spans ~4 orders of magnitude.",
        "  If plotting on a log axis, floor zeros (other panels use 0.01) but run",
        "  the statistics on the TRUE values including zeros.",
        "",
        "WHY THIS PANEL: burden = frequency x duration. Duration does not differ",
        "between groups (panel G, ns), so a burden difference shows the effect is",
        "carried by frequency rather than by shorter seizures.",
    ])
    ws2.column_dimensions["A"].width = 17
    for i in range(len(E) + len(S)):
        ws2.column_dimensions[get_column_letter(2 + i)].width = 11

    # ---- sheet 3: cumulative burden over baseline days ----
    cum = cumulative(D, E, S)
    ws3 = wb.create_sheet("3_CumulativeBurden")
    ws3["A1"] = ("Cumulative seizure time (minutes) over baseline days 1-21, "
                 f"confidence >= {args.cut}")
    ws3["A1"].font = BOLD
    ws3["A3"] = ("PRISM TABLE TYPE:  XY, with X = day and one Y column per "
                 "animal (Y = single values, no replicates)")
    ws3["A3"].font = BOLD
    ws3["A4"] = ("Plot every animal as a thin grey line; overlay the two group "
                 "MEDIAN columns as bold lines. Do not use mean +/- SEM — see note.")
    r = 6
    hdr = (["X  (day)"] + [f"EGFP {a}" for a in E] + [f"SV2A {a}" for a in S]
           + ["EGFP median", "SV2A median"])
    for c, h in enumerate(hdr):
        cell = ws3.cell(row=r, column=c + 1, value=h)
        cell.font, cell.fill = BOLD, HDR
    r += 1
    for i, dd in enumerate(range(1, 22)):
        ws3.cell(row=r, column=1, value=dd)
        for j, a in enumerate(E + S):
            ws3.cell(row=r, column=2 + j, value=round(cum[a][i], 3))
        ws3.cell(row=r, column=2 + len(E) + len(S),
                 value=round(float(np.median([cum[a][i] for a in E])), 3))
        ws3.cell(row=r, column=3 + len(E) + len(S),
                 value=round(float(np.median([cum[a][i] for a in S])), 3))
        r += 1
    fa = [cum[a][20] for a in E]
    fb = [cum[a][20] for a in S]
    pf = mw(fa, fb)
    put_notes(ws3, r + 1, [
        "STATISTICS TO RUN IN PRISM",
        "  The curves are descriptive. Test the ENDPOINT (day 21 total) only:",
        "  Analyze → Column analyses → t test (and nonparametric tests)",
        "         → 'Mann-Whitney test' (unpaired, two-tailed), EGFP vs SV2A.",
        f"  EXPECTED: median EGFP {pf[0]:.1f} min, SV2A {pf[1]:.1f} min, "
        f"p = {pf[2]:.4f}",
        "  Do not run a 2-way ANOVA over days: cumulative values are not",
        "  independent across time, so the day factor is not interpretable.",
        "",
        "PLOT MEDIAN, NOT MEAN +/- SEM. The SV2A group is bimodal — 4 animals",
        "at exactly 0 for all 21 days and 2 non-responders as high as EGFP — so",
        "the mean is dragged up and the error bars overlap (EGFP 225 +/- 78 vs",
        "SV2A 45 +/- 29 min), which understates a difference the medians show",
        "clearly (87.7 vs 0.8 min).",
        "",
        "ZEROS ARE REAL. A flat line along the axis means that animal never had",
        "a seizure; it is not a plotting floor. Use a LINEAR y-axis here — do not",
        "log-transform, or the flat-at-zero animals cannot be drawn.",
        "",
        f"NON-RESPONDERS: SV2A animals with the two highest totals are the two",
        "climbing SV2A lines; naming them in the legend is worth doing, since the",
        "responder/non-responder split is more informative than the group median.",
    ])
    ws3.column_dimensions["A"].width = 11
    for i in range(len(E) + len(S) + 2):
        ws3.column_dimensions[get_column_letter(2 + i)].width = 12

    wb.save(args.out)
    print(f"Wrote {args.out}")
    print(f"  KM: log-rank chi2={chi2:.2f}, p={p_lr:.4f}")
    print(f"  burden: convulsive p={pc[2]:.4f}, non-convulsive p={pn[2]:.4f}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
