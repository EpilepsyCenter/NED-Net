#!/usr/bin/env python
"""Build the Source Data workbook for the SV2A manuscript.

ONE SHEET PER FIGURE, with the panels stacked down the sheet — a workbook with
a sheet per panel runs to 40+ tabs once every figure is included.

Panel-to-table mapping was established from the manuscript's own figure
citations and from the figure PDFs, not guessed. Values are read from the Prism
analysis files so the numbers are exactly what is plotted; Fig 3f and 3g are the
exception, having been added after the analysis file was built, and are
recomputed from the same database and definitions.

Figures NOT covered here, because their data live outside the Prism files:
    Supplementary Figure 1 (NED-Net validation)

    python scripts/paper_stats/make_source_data.py
"""
from __future__ import annotations

import argparse
import csv
import io
import json
import os
import sys
import zipfile

import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from prism_rewrite_all import Data                      # noqa: E402
from fig3_replacement_panels import cumulative          # noqa: E402

BOLD = Font(bold=True)
BIG = Font(bold=True, size=12)
ITAL = Font(italic=True, size=9)
HDR = PatternFill("solid", fgColor="DDDDDD")
BASE = {1, 2, 3}

ANALYSIS = ("/Users/marcoledri/Dropbox/Work/Manuscripts and papers/"
            "SV2A paper/Data analysis")
SEIZ = "Seizure_data_conf0.5.prism"
IMMU = "SV2A Immuno.prism"
BEHAV = "All Behaviour tests.prism"
EPHYS = "Ephys/Ephys_summary.prism"
ADSPK = "AD/AD_Interictal_spikes.prism"
OFTB = "AD/5xFAD_SV2A_GT_OFT baseline.prism"
OFT2 = "AD/5xFAD_SV2A_GT_OFT repeat 2 + comparisons.prism"
BARN = "AD/5xFAD_SV2A_GT_Barnes repeat 2 + comparisons.prism"
NORT = "AD/5xFAD_SV2A_GT_NORT repeat2 + comparisons.prism"
OLT2 = "AD/5xFAD_SV2A_GT_OLT repeat2 + comparisons.prism"

AD_FILES = None   # set after the file constants below; see group_labels()


def group_labels(book, width):
    """Column headings for a table, by which cohort the file belongs to.

    The 5xFAD experiments use WT / 5xFAD-EGFP / 5xFAD-SV2A, not the SV2A
    cohort's EGFP / SV2A, and the two-group baseline panels are WT vs 5xFAD.
    """
    if _is_5x(book):
        names = ["WT", "5xFAD-EGFP", "5xFAD-SV2A"]
        if width == 2:
            names = ["WT", "5xFAD"]
    else:
        names = ["EGFP", "SV2A"]
    if width <= len(names):
        return names[:width]
    return names + [f"col{i+1}" for i in range(len(names), width)]


def block_note(book):
    return ("Each column is one animal; the header row gives its group. "
            "Blank = no value for that animal.")


def block_split(book, width):
    """-> (n_label_cols, block_width) for a Prism grouped table.

    Blocks are equal width, one per group, so the data columns must divide by
    the number of groups. Some tables carry a leading label column (trial
    number, condition) and some do not — assuming one where there is none
    shifts every block and silently drops values off the end, so pick whichever
    divides exactly, preferring a label column when both do.
    """
    n = len(group_labels(book, 3 if _is_5x(book) else 2))
    for n_hdr in (1, 0):
        if width > n_hdr and (width - n_hdr) % n == 0:
            return n_hdr, (width - n_hdr) // n
    return 1, max(1, (width - 1) // n)


def block_header(book, width):
    """Column headings naming the group of every animal column."""
    names = group_labels(book, 3 if _is_5x(book) else 2)
    n_hdr, per = block_split(book, width)
    head = [""] * n_hdr
    for g in names:
        for i in range(per):
            head.append(f"{g} {i + 1}")
    while len(head) < width:
        head.append("")
    return head[:width]


def _is_5x(book):
    return bool(book) and book.startswith("AD/") and "Interictal" not in book


MW = "Mann-Whitney U, two-tailed"
WSR = "Wilcoxon matched-pairs signed-rank (within animal)"
KW = "Kruskal-Wallis with Dunn's multiple comparisons"
AN = "Two-way ANOVA with Dunnett's multiple comparisons"

# (figure, panel, prism file, table title, caption, test, layout)
PANELS = [
    ("Figure 1", "j", IMMU, "Cell density (individual)",
     "SV2A-positive cell density (cells/mm2) by hippocampal region", MW, "ROWBLOCK"),
    ("Figure 1", "k", IMMU, "MGV (individual)",
     "SV2A immunoreactivity, mean grey value, by hippocampal region", MW, "ROWBLOCK"),

    # Figure 2 — in vitro electrophysiology. sIPSC = "IPSC", mIPSC = "TTX"
    # (miniature currents recorded in tetrodotoxin).
    ("Figure 2", "c", EPHYS, "Frequency grouped",
     "sIPSC / mIPSC frequency (Hz)", MW, "ROWBLOCK"),
    ("Figure 2", "d", EPHYS, "IPSC IEI distribution",
     "sIPSC inter-event interval, cumulative distribution",
     "Kolmogorov-Smirnov", "RAW"),
    ("Figure 2", "e", EPHYS, "TTX Frequency distribution",
     "mIPSC inter-event interval, cumulative distribution",
     "Kolmogorov-Smirnov", "RAW"),
    ("Figure 2", "f", EPHYS, "Amplitude grouped",
     "sIPSC / mIPSC amplitude (pA)", MW, "ROWBLOCK"),
    ("Figure 2", "g", EPHYS, "IPSC Amplitude distribution",
     "sIPSC amplitude, cumulative distribution", "Kolmogorov-Smirnov", "RAW"),
    ("Figure 2", "h", EPHYS, "TTX Amplitude distribution",
     "mIPSC amplitude, cumulative distribution", "Kolmogorov-Smirnov", "RAW"),
    ("Figure 2", "i", EPHYS, "RT grouped",
     "sIPSC / mIPSC rise time (ms)", MW, "ROWBLOCK"),
    ("Figure 2", "j", EPHYS, "IPSC RT distribution",
     "sIPSC rise time, cumulative distribution", "Kolmogorov-Smirnov", "RAW"),
    ("Figure 2", "k", EPHYS, "TTX RT distribution",
     "mIPSC rise time, cumulative distribution", "Kolmogorov-Smirnov", "RAW"),

    ("Figure 3", "c", SEIZ, "Seizures/day",
     "Seizures per 24 h of recording, baseline (weeks 1-3)", MW, "GG"),
    ("Figure 3", "d", SEIZ, "Convulsive free days baseline",
     "% of recorded days without a convulsive seizure, baseline", MW, "COL2"),
    ("Figure 3", "e", SEIZ, "Per week seizures",
     "All seizures per 24 h, by baseline week", "Two-way ANOVA (week x group)", "WEEK"),
    ("Figure 3", "f", None, None,
     "Time to first convulsive seizure, baseline", "Log-rank (Mantel-Cox)", "SURV"),
    ("Figure 3", "g", None, None,
     "Cumulative convulsive seizure burden (min), baseline days 1-21",
     "Mann-Whitney U on the day-21 endpoint", "CUM"),
    ("Figure 3", "h", SEIZ, "Duration",
     "Mean seizure duration (s), whole recording", MW, "GG"),
    ("Figure 3", "i", SEIZ, "Average SE seizure grade",
     "Average behavioural seizure grade during status epilepticus", MW, "COL2"),
    ("Figure 3", "j", SEIZ, "Cumulative SE seizure grade",
     "Cumulative behavioural seizure grade during SE", MW, "COL2"),
    ("Figure 3", "k", SEIZ, "SE correlation",
     "Cumulative SE grade vs chronic seizure rate", "Linear regression per group", "RAW"),

    ("Figure 4", "c", SEIZ, "IS/hour",
     "Interictal spikes per animal-hour, baseline", MW, "COL2"),
    ("Figure 4", "d", SEIZ, "IS duration",
     "Mean interictal spike duration (s), baseline", MW, "COL2"),
    ("Figure 4", "e", SEIZ, "ISI frequency dist",
     "Inter-spike-interval probability density (ISIs pooled within group)",
     "Descriptive", "RAW"),
    ("Figure 4", "f", SEIZ, "IS ISI cumulative",
     "Inter-spike-interval cumulative distribution (pooled within group)",
     "Kolmogorov-Smirnov on pooled ISIs", "RAW"),

    ("Figure 5", "a", SEIZ, "Events/animal-hour ALL DAYS",
     "All seizures per animal-hour, by recording day", "Descriptive (mean +/- s.e.m.)", "DAY"),
    ("Figure 5", "b", SEIZ, "Convulsive free days pre-post LEV",
     "% days convulsive-free, baseline vs levetiracetam", WSR, "PP"),
    ("Figure 5", "c", SEIZ, "Convulsive pre-post LEV",
     "Convulsive seizures per 24 h, baseline vs levetiracetam", WSR, "PP"),
    ("Figure 5", "d", SEIZ, "Non-Convulsive pre-post LEV",
     "Non-convulsive seizures per 24 h, baseline vs levetiracetam", WSR, "PP"),
    ("Figure 5", "e", SEIZ, "Convulsive Duration pre-post LEV",
     "Mean convulsive seizure duration (s), baseline vs levetiracetam", WSR, "PP"),
    ("Figure 5", "f", SEIZ, "Non-Convulsive Duration pre-post LEV",
     "Mean non-convulsive seizure duration (s), baseline vs levetiracetam", WSR, "PP"),
    ("Figure 5", "g", SEIZ, "IS/hour pre-post LEV",
     "Interictal spikes per animal-hour, baseline vs levetiracetam", WSR, "PP"),
    ("Figure 5", "h", SEIZ, "IS ISI cumulative pre-post LEV EGFP",
     "ISI cumulative distribution, EGFP, baseline vs LEV",
     "Kolmogorov-Smirnov on pooled ISIs", "RAW"),
    ("Figure 5", "h", SEIZ, "IS ISI cumulative pre-post LEV SV2A",
     "ISI cumulative distribution, SV2A, baseline vs LEV",
     "Kolmogorov-Smirnov on pooled ISIs", "RAW"),
    ("Figure 5", "i", SEIZ, "IS duration pre-post LEV",
     "Mean spike duration (s), baseline vs levetiracetam", WSR, "PP"),

    ("Figure 6", "b", BEHAV, "OLT time grouped",
     "Object location test — exploration time (s)", MW, "ROWBLOCK"),
    ("Figure 6", "c", BEHAV, "OLT Head entries grouped",
     "Object location test — head entries", MW, "ROWBLOCK"),
    ("Figure 6", "d", BEHAV,
     "*Descrimination ratio between Object A and object D Time Object location test trial",
     "Object location test — discrimination index", MW, "AUTO"),
    ("Figure 6", "f", BEHAV, "NOR Time grouped",
     "Novel object recognition — exploration time (s)", MW, "ROWBLOCK"),
    ("Figure 6", "g", BEHAV, "NOR Head entries grouped",
     "Novel object recognition — head entries", MW, "ROWBLOCK"),
    ("Figure 6", "h", BEHAV,
     "*DESCRIMINATION Between NOVAL And familliar time NOR",
     "Novel object recognition — discrimination index", MW, "AUTO"),
    ("Figure 6", "j", BEHAV, "Nest building",
     "Nest-building score (Deacon protocol)", MW, "AUTO"),
    ("Figure 6", "l", BEHAV, "Time Mobile OFT",
     "Open field — time mobile (s)", MW, "AUTO"),
    ("Figure 6", "m", BEHAV, "Center arena enteries OFT",
     "Open field — centre entries", MW, "AUTO"),
    ("Figure 6", "n", BEHAV, "Time in Center Arena OFT",
     "Open field — time in centre (s)", MW, "AUTO"),
    ("Figure 6", "o", BEHAV, "% Time in Center Arena OFT",
     "Open field — time in centre (% of mobile time)", MW, "AUTO"),

    # Figure 7 — 5xFAD open field and Barnes maze. Groups: WT, 5xFAD-EGFP,
    # 5xFAD-SV2A (panel a is the WT vs 5xFAD baseline comparison, two groups).
    ("Figure 7", "a", OFTB, "OFT: distance",
     "Open field distance (m), WT vs 5xFAD at baseline (2 columns)",
     "Unpaired t-test", "AUTO"),
    ("Figure 7", "b", OFT2, "Distance",
     "Open field distance (m) after treatment "
     "(columns: WT, 5xFAD-EGFP, 5xFAD-SV2A)", "Kruskal-Wallis with Dunn's multiple comparisons", "AUTO"),
    ("Figure 7", "c", OFT2, "Centre time ratio",
     "Time in centre (fraction of total), pre vs post treatment "
     "(rows Pre/Post; columns are animals, WT then 5xFAD-EGFP then 5xFAD-SV2A)",
     "Kruskal-Wallis with Dunn's multiple comparisons", "ROWBLOCK"),
    ("Figure 7", "d", BARN, "delta AUC: primary latency",
     "Barnes maze, change in AUC of primary latency", "Kruskal-Wallis with Dunn's multiple comparisons", "AUTO"),
    ("Figure 7", "e", BARN, "delta AUC: total latency",
     "Barnes maze, change in AUC of total latency", "Kruskal-Wallis with Dunn's multiple comparisons", "AUTO"),
    ("Figure 7", "f", BARN, " delta AUC: total errors",
     "Barnes maze, change in AUC of total errors", "Kruskal-Wallis with Dunn's multiple comparisons", "AUTO"),

    # Extended Data 2 — object location and novel object recognition, 5xFAD triad.
    ("Extended Data 2", "a", OLT2, "Time in object (test) GROUPED",
     "Object location: exploration time (s), familiar vs displaced", KW, "ROWBLOCK"),
    ("Extended Data 2", "b", OLT2, "zone entries (no. of investigations) GROUPED",
     "Object location: head entries, familiar vs displaced", KW, "ROWBLOCK"),
    ("Extended Data 2", "c", OLT2, "DI test NEW",
     "Object location: discrimination index "
     "(columns: WT-Sham, 5xFAD-EGFP, 5xFAD-SV2A)", KW, "AUTO"),
    ("Extended Data 2", "d", NORT, "Time in object test GROUPED",
     "Novel object: exploration time (s), familiar vs novel", KW, "ROWBLOCK"),
    ("Extended Data 2", "e", NORT, "Head entries test GROUPED",
     "Novel object: head entries, familiar vs novel", KW, "ROWBLOCK"),
    ("Extended Data 2", "f", NORT, "Data 14",
     "Novel object: discrimination index "
     "(columns: WT-Sham, 5xFAD-EGFP, 5xFAD-SV2A)", KW, "AUTO"),

    # Extended Data 3 — Barnes maze training curves and probe trial.
    ("Extended Data 3", "a", BARN, "Primary latency, training",
     "Barnes maze training: primary latency (s), rows = trials 1-5", AN, "ROWBLOCK"),
    ("Extended Data 3", "b", BARN, "Total latency, training",
     "Barnes maze training: total latency (s), rows = trials 1-5", AN, "ROWBLOCK"),
    ("Extended Data 3", "c", BARN, "Primary errors, training",
     "Barnes maze training: primary errors, rows = trials 1-5", AN, "ROWBLOCK"),
    ("Extended Data 3", "d", BARN, "Total errors, training",
     "Barnes maze training: total errors, rows = trials 1-5", AN, "ROWBLOCK"),
    ("Extended Data 3", "e", BARN, "Latency, probe",
     "Barnes maze probe trial: latency (s)", KW, "AUTO"),
    ("Extended Data 3", "f", BARN, "Errors, probe",
     "Barnes maze probe trial: errors", KW, "AUTO"),

    # AD (5xFAD) cohort = Extended Data Fig. 1 per the manuscript and the
    # Extended Data legends. The PDF file is named "...Figure 3" — the figure
    # FILES are shuffled relative to the text; see README.
    ("Extended Data 1 (AD)", "c", ADSPK, "spike_rate_by_day",
     "Interictal spikes per animal-hour, by recording day (col 1 = day, then "
     "animals 1-4, then cohort mean)", "Linear regression", "RAW"),
    ("Extended Data 1 (AD)", "d", ADSPK, "spike_rate_by_week",
     "Interictal spikes per animal-hour, by week (rows = animals 1-4)",
     "Descriptive", "RAW"),
    ("Extended Data 1 (AD)", "-", None, None,
     "Seizures detected in the AD cohort (not plotted; quoted in the text)",
     "Descriptive", "ADSEIZ"),
]


def _num(c):
    c = (c or "").strip()
    if c in ("", "-"):
        return None
    try:
        return float(c.replace(",", "."))
    except ValueError:
        return None


def read_tables(path):
    z = zipfile.ZipFile(path)
    titles = {}
    for n in z.namelist():
        if n.endswith("sheet.json"):
            d = json.load(io.TextIOWrapper(z.open(n)))
            if (d.get("@class") == "DataSheet"
                    and "ANALYSIS_VIEW" not in (d.get("flags") or [])):
                tu = (d.get("table") or {}).get("uid")
                if tu:
                    titles.setdefault(d.get("title", ""), tu)
    out = {}
    for t, tu in titles.items():
        try:
            rows = list(csv.reader(io.TextIOWrapper(
                z.open(f"data/tables/{tu}/data.csv"))))
        except KeyError:
            continue
        while rows and not any((c or "").strip() for c in rows[-1]):
            rows.pop()
        if rows:
            out[t] = rows
    return out


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--out", default="SourceData_SV2A.xlsx")
    ap.add_argument("--cut", type=float, default=0.5)
    args = ap.parse_args()

    books = {}
    for f in (SEIZ, IMMU, BEHAV, EPHYS, ADSPK, OFTB, OFT2, BARN, NORT, OLT2):
        p = f if os.path.exists(f) else os.path.join(ANALYSIS, f)
        if not os.path.exists(p):
            print(f"  !! missing {f}")
            continue
        books[f] = read_tables(p)

    D = Data(args.cut)
    E, S = D.egfp, D.sv2a
    nE, nS = len(E), len(S)
    ids = [f"EGFP {a}" for a in E] + [f"SV2A {a}" for a in S]
    groups = ["EGFP"] * nE + ["SV2A"] * nS

    wb = openpyxl.Workbook()
    readme = wb.active
    readme.title = "README"

    figures = []
    for fig, panel, book, title, caption, test, layout in PANELS:
        if fig not in figures:
            figures.append(fig)
    sheets = {f: wb.create_sheet(f) for f in figures}
    rowptr = {f: 1 for f in figures}
    covered = []

    def emit(sh, f, text, font=None, cols=None):
        r = rowptr[f]
        if cols is None:
            sh.cell(row=r, column=1, value=text).font = font or ITAL
        else:
            for c, v in enumerate(cols):
                cell = sh.cell(row=r, column=c + 1, value=v)
                if font:
                    cell.font = font
                if font is BOLD:
                    cell.fill = HDR
        rowptr[f] = r + 1
        return r

    for fig, panel, book, title, caption, test, layout in PANELS:
        sh = sheets[fig]
        rows = None
        if book:
            rows = books.get(book, {}).get(title)
            if rows is None:
                print(f"  !! {fig}{panel}: '{title}' not found in {book}")
                continue
        emit(sh, fig, f"{fig}{panel} — {caption}", BIG)
        emit(sh, fig, f"Statistical test: {test}", ITAL)

        if layout == "GG":
            width = max(len(r) for r in rows)
            b1, b2 = 1, 1 + (width - 1) // 2
            labs = [(rows[i][0] or f"row{i}") for i in range(min(2, len(rows)))]
            emit(sh, fig, None, BOLD, ["animal", "group"] + labs)
            for k in range(nE + nS):
                base = b1 + k if k < nE else b2 + (k - nE)
                vals = [_num(rows[j][base]) if base < len(rows[j]) else None
                        for j in range(len(labs))]
                emit(sh, fig, None, None, [ids[k], groups[k]] + vals)

        elif layout == "PP":
            width = max(len(r) for r in rows)
            b1, b2 = 1, 1 + (width - 1) // 2
            emit(sh, fig, None, BOLD,
                 ["animal", "group", "baseline", "levetiracetam"])
            for k in range(nE + nS):
                ri = 0 if k < nE else 1
                off = k if k < nE else k - nE
                vals = []
                for b in (b1, b2):
                    c = b + off
                    vals.append(_num(rows[ri][c])
                                if ri < len(rows) and c < len(rows[ri]) else None)
                emit(sh, fig, None, None, [ids[k], groups[k]] + vals)

        elif layout == "COL2":
            emit(sh, fig, None, BOLD, ["animal", "group", "value"])
            k = 0
            for cj, gname in ((0, "EGFP"), (1, "SV2A")):
                for r_ in rows:
                    v = _num(r_[cj]) if cj < len(r_) else None
                    if v is None:
                        continue
                    emit(sh, fig, None, None,
                         [ids[k] if k < len(ids) else "", gname, v])
                    k += 1

        elif layout in ("WEEK", "DAY"):
            width = max(len(r) for r in rows)
            b1, b2 = 1, 1 + (width - 1) // 2
            emit(sh, fig, None, BOLD,
                 ["week" if layout == "WEEK" else "day"] + ids)
            for r_ in rows:
                idx = _num(r_[0]) if r_ else None
                if idx is None:
                    continue
                vals = []
                for k in range(nE + nS):
                    c = b1 + k if k < nE else b2 + (k - nE)
                    vals.append(_num(r_[c]) if c < len(r_) else None)
                emit(sh, fig, None, None, [int(idx)] + vals)

        elif layout == "ROWBLOCK":
            width = max(len(r) for r in rows)
            n_hdr, _per = block_split(book, width)
            emit(sh, fig, block_note(book), ITAL)
            emit(sh, fig, None, BOLD, block_header(book, width))
            for r_ in rows:
                if not any((c or "").strip() for c in r_):
                    continue
                emit(sh, fig, None, None,
                     [(r_[i] or "") if i < n_hdr else _num(r_[i])
                      for i in range(len(r_))])

        elif layout == "AUTO":
            width = max(len(r) for r in rows)
            if width <= 3:
                emit(sh, fig, None, BOLD, group_labels(book, width))
                for r_ in rows:
                    vals = [_num(c) for c in r_]
                    if any(v is not None for v in vals):
                        emit(sh, fig, None, None, vals)
            else:
                n_hdr, _per = block_split(book, width)
                emit(sh, fig, block_note(book), ITAL)
                emit(sh, fig, None, BOLD, block_header(book, width))
                for r_ in rows:
                    if any((c or "").strip() for c in r_):
                        emit(sh, fig, None, None,
                             [(r_[i] or "") if i < n_hdr else _num(r_[i])
                              for i in range(len(r_))])

        elif layout == "SURV":
            emit(sh, fig,
                 "1 = convulsive seizure that day; 0 = censored (none during "
                 "baseline; day = last recorded day).", ITAL)
            emit(sh, fig, None, BOLD,
                 ["animal", "group", "day", "event (1) / censored (0)"])
            first = {}
            for ci, a, t, _d in D.ev:
                if t == "convulsive" and D.wk.get(ci) in BASE:
                    dn = D.day.get(ci)
                    if dn and (a not in first or dn < first[a]):
                        first[a] = dn
            for k, a in enumerate(E + S):
                days = [D.day[ci] for ci, aa, _s in D.fa
                        if aa == a and D.wk.get(ci) in BASE]
                t, e = ((first[a], 1) if a in first
                        else (max(days) if days else 21, 0))
                emit(sh, fig, None, None, [ids[k], groups[k], t, e])

        elif layout == "CUM":
            cum = cumulative(D, E, S, "convulsive")
            emit(sh, fig, None, BOLD,
                 ["day"] + ids + ["EGFP median", "SV2A median"])
            for i, dd in enumerate(range(1, 22)):
                vals = [round(cum[a][i], 3) for a in E + S]
                emit(sh, fig, None, None, [dd] + vals
                     + [round(float(np.median([cum[a][i] for a in E])), 3),
                        round(float(np.median([cum[a][i] for a in S])), 3)])

        elif layout == "ADSEIZ":
            import sqlite3
            adb = os.path.expanduser(
                "~/.eeg_seizure_analyzer/projects/ad_seizures_local.db")
            if not os.path.exists(adb):
                emit(sh, fig, f"(not built: {adb} not found)", ITAL)
            else:
                emit(sh, fig,
                     "Confidence >= 0.5; every event visually confirmed. "
                     "Recording 462 h per animal (1,847 animal-hours).", ITAL)
                emit(sh, fig, None, BOLD,
                     ["animal", "week", "convulsive", "non-convulsive"])
                con = sqlite3.connect(adb)
                rowsq = con.execute(
                    "SELECT e.animal_id, "
                    "  CASE WHEN c.path LIKE '%Week_1%' THEN 1 "
                    "       WHEN c.path LIKE '%Week_2%' THEN 2 ELSE 3 END wk, "
                    "  SUM(e.type='convulsive'), SUM(e.type='non_convulsive') "
                    "FROM events e JOIN chunks c ON e.chunk_id=c.id "
                    "WHERE e.cnn_confidence>=0.5 "
                    "GROUP BY e.animal_id, wk ORDER BY e.animal_id, wk").fetchall()
                con.close()
                for a, w, cv, nc in rowsq:
                    emit(sh, fig, None, None, [f"AD animal {a}", w, cv, nc])

        else:   # RAW
            emit(sh, fig, "Curve data as plotted; columns as in the figure.", ITAL)
            for r_ in rows:
                vals = [_num(c) if _num(c) is not None else (c or None) for c in r_]
                if any(v is not None for v in vals):
                    emit(sh, fig, None, None, vals)

        rowptr[fig] += 2
        covered.append((fig, panel, caption))

    for f, sh in sheets.items():
        sh.column_dimensions["A"].width = 20
        sh.column_dimensions["B"].width = 12
        for i in range(2, 30):
            sh.column_dimensions[get_column_letter(i + 1)].width = 12

    lines = [
        ("Source Data — SV2A manuscript", BIG),
        ("", ITAL),
        ("One sheet per figure; panels are stacked down each sheet in order.", ITAL),
        ("", ITAL),
        ("SEIZURE AND SPIKE DATA (Figures 3-5)", BOLD),
        ("  Detection database: lunarc_detect_wk1-6_final.db (NED-Net)", ITAL),
        (f"  Detector-confidence threshold >= {args.cut}", ITAL),
        ("  Baseline = recording weeks 1-3; levetiracetam = weeks 4-5.", ITAL),
        ("  Seizure rates per 24 h of recording; spike rates per animal-hour.", ITAL),
        (f"  n = {nE} EGFP / {nS} SV2A; animals 355676, 372837, 30 and the", ITAL),
        ("  unassigned channel 'x' are excluded throughout.", ITAL),
        ("  Animal order used in every sheet:", ITAL),
        ("    EGFP: " + ", ".join(E), ITAL),
        ("    SV2A: " + ", ".join(S), ITAL),
        ("  Fig 3f and 3g were added after the analysis file was built and are", ITAL),
        ("  recomputed from the same database and definitions.", ITAL),
        ("", ITAL),
        ("HISTOLOGY (Figure 1j,k) — from 'SV2A Immuno.prism'", BOLD),
        ("BEHAVIOUR (Figure 6) — from 'All Behaviour tests.prism'", BOLD),
        ("  For these, values are individual animals in the order plotted,", ITAL),
        ("  EGFP block followed by SV2A block; animal identifiers were not", ITAL),
        ("  recorded in the analysis files.", ITAL),
        ("", ITAL),
        ("ELECTROPHYSIOLOGY (Figure 2) — from 'Ephys/Ephys_summary.prism'", BOLD),
        ("  sIPSC = tables named 'IPSC'; mIPSC = tables named 'TTX'", ITAL),
        ("  (miniature currents recorded in tetrodotoxin).", ITAL),
        ("  CHECK: the current Figure 2 PDF labels panel e's x-axis", ITAL),
        ("  'Amplitude (pA)', but the manuscript describes panels d,e as", ITAL),
        ("  inter-event interval. Panels here follow the manuscript.", ITAL),
        ("", ITAL),
        ("AD COHORT (5xFAD) — from 'AD/AD_Interictal_spikes.prism'", BOLD),
        ("  Spike rates use the standard baseline-relative detector.", ITAL),
        ("  The seizure block is not plotted; it supports the numbers quoted", ITAL),
        ("  in the text, and every event was visually confirmed.", ITAL),
        ("  Numbered Extended Data 1, following the manuscript and the", ITAL),
        ("  Extended Data legends.", ITAL),
        ("", ITAL),
        ("  CHECK — THE EXTENDED DATA PDF FILES ARE MISNAMED.", BOLD),
        ("  Manuscript text and legends say:  ED1 = AD interictal spikes,", ITAL),
        ("     ED2 = object location / novel object, ED3 = Barnes maze.", ITAL),
        ("  The PDF files currently hold:  'ExtendedData_Figure 1.pdf' = OLT/NOR,", ITAL),
        ("     'ExtendedData_Figure 2.pdf' = Barnes, 'ExtendedData_Figure 3.pdf'", ITAL),
        ("     = AD spikes. The three files need renaming 3->1, 1->2, 2->3", ITAL),
        ("     before submission, or the citations will point at the wrong", ITAL),
        ("     figures.", ITAL),
        ("", ITAL),
        ("5xFAD COHORT (Figure 7, Extended Data 2 and 3)", BOLD),
        ("  From the 'AD' folder: OFT baseline and repeat 2, Barnes repeat 2,", ITAL),
        ("  NORT repeat 2 and OLT repeat 2 Prism files.", ITAL),
        ("  Groups are WT (or WT-Sham/WT-Ctrl), 5xFAD-EGFP and 5xFAD-SV2A,", ITAL),
        ("  in that column order. Fig 7a is the two-group WT vs 5xFAD baseline.", ITAL),
        ("  n = 10 WT-Ctrl, 5 5xFAD-EGFP, 9 5xFAD-SV2A in the Barnes maze; two", ITAL),
        ("  5xFAD-EGFP animals were lost between the recognition tasks and the", ITAL),
        ("  Barnes maze.", ITAL),
        ("", ITAL),
        ("NOT INCLUDED", BOLD),
        ("  Supplementary Figure 1 (NED-Net validation).", ITAL),
        ("", ITAL),
        ("CONTENTS", BOLD),
    ]
    r = 1
    for text, font in lines:
        readme.cell(row=r, column=1, value=text).font = font
        r += 1
    last = None
    for fig, panel, caption in covered:
        if fig != last:
            readme.cell(row=r, column=1, value=f"  {fig}").font = BOLD
            r += 1
            last = fig
        readme.cell(row=r, column=1, value=f"      {panel}   {caption}").font = ITAL
        r += 1
    readme.column_dimensions["A"].width = 100

    wb.save(args.out)
    print(f"Wrote {args.out}: {len(figures)} figure sheets, "
          f"{len(covered)} panels + README")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
